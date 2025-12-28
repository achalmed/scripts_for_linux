#!/usr/bin/env python3
"""
PDF Page Counter - Contador de páginas de archivos PDF
Autor: Edison Achalma
Universidad Nacional de San Cristóbal de Huamanga

v1.1.0
"""

import os
import sys
from pathlib import Path
from typing import List, Tuple, Dict
import argparse
from datetime import datetime

try:
    from PyPDF2 import PdfReader
except ImportError:
    print("❌ Error: PyPDF2 no está instalado.")
    print("Por favor, instala con: conda install -c conda-forge pypdf2")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("❌ Error: openpyxl no está instalado.")
    print("Por favor, instala con: conda install -c conda-forge openpyxl")
    sys.exit(1)


# ========================================================================
# CONFIGURACIÓN - Personaliza estas rutas según tu sistema
# ========================================================================

# Ruta base donde están todos tus blogs
RUTA_BASE_PUBLICACIONES = "/home/achalmaedison/Documents/publicaciones"

# Lista de blogs a procesar (estructura estándar con _site)
BLOGS_ESTANDAR = [
    "actus-mercator",
    "aequilibria",
    "axiomata",
    "chaska",
    "dialectica-y-mercado",
    "epsilon-y-beta",
    "methodica",
    "numerus-scriptum",
    "optimums",
    "pecunia-fluxus",
    "res-publica",
]

# Blogs dentro de website-achalma (no tienen _site propio, están en _site/blog o _site/teching)
BLOGS_WEBSITE_ACHALMA = {
    "blog": "website-achalma/_site/blog",
    "teching": "website-achalma/_site/teching",
}

# Directorio donde se guardarán los reportes Excel
DIRECTORIO_EXCEL = "excel_databases"

# ========================================================================
# FUNCIONES AUXILIARES
# ========================================================================

def imprimir_encabezado():
    """Imprime un encabezado bonito para el script"""
    print("\n" + "=" * 80)
    print("📊 PDF PAGE COUNTER - CONTADOR DE PÁGINAS PDF".center(80))
    print("=" * 80)
    print("👤 Autor: Edison Achalma")
    print("🏛️  Universidad Nacional de San Cristóbal de Huamanga")
    print("📅 Fecha:", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    print("=" * 80 + "\n")


def imprimir_seccion(titulo: str):
    """Imprime una sección con formato"""
    print("\n" + "─" * 80)
    print(f"📌 {titulo}")
    print("─" * 80)


def imprimir_resumen(total_archivos: int, total_paginas: int, errores: int):
    """Imprime un resumen final bonito"""
    print("\n" + "=" * 80)
    print("📈 RESUMEN FINAL".center(80))
    print("=" * 80)
    print(f"✅ Archivos procesados exitosamente: {total_archivos - errores}")
    print(f"❌ Archivos con errores: {errores}")
    print(f"📄 Total de archivos analizados: {total_archivos}")
    print(f"📑 Total de páginas contadas: {total_paginas:,}")
    print("=" * 80 + "\n")


def crear_directorio_excel():
    """Crea el directorio para guardar los archivos Excel si no existe"""
    script_dir = Path(__file__).parent
    excel_dir = script_dir / DIRECTORIO_EXCEL
    excel_dir.mkdir(exist_ok=True)
    return excel_dir


def contar_paginas_pdf(ruta_pdf: str) -> int:
    """
    Cuenta el número de páginas de un archivo PDF.
    
    Args:
        ruta_pdf: Ruta al archivo PDF
        
    Returns:
        Número de páginas del PDF, o -1 si hay error
    """
    try:
        reader = PdfReader(ruta_pdf)
        return len(reader.pages)
    except Exception as e:
        return -1


def buscar_pdfs_en_directorio(directorio: str, solo_index: bool = True) -> List[Tuple[str, int, str]]:
    """
    Busca archivos PDF recursivamente en un directorio.
    
    Args:
        directorio: Directorio raíz donde buscar
        solo_index: Si es True, solo busca archivos llamados 'index.pdf'
        
    Returns:
        Lista de tuplas (ruta_relativa, numero_paginas, estado)
    """
    resultados = []
    directorio_path = Path(directorio)
    
    if not directorio_path.exists():
        print(f"⚠️  Directorio no encontrado: {directorio}")
        return resultados
    
    # Patrón de búsqueda
    patron = "**/index.pdf" if solo_index else "**/*.pdf"
    
    # Buscar archivos
    archivos_encontrados = list(directorio_path.glob(patron))
    
    if not archivos_encontrados:
        print(f"   ℹ️  No se encontraron archivos en: {directorio_path.name}")
        return resultados
    
    for pdf_path in archivos_encontrados:
        if pdf_path.is_file():
            ruta_relativa = str(pdf_path.relative_to(directorio_path))
            paginas = contar_paginas_pdf(str(pdf_path))
            
            if paginas > 0:
                resultados.append((ruta_relativa, paginas, "OK"))
                print(f"   ✓ {ruta_relativa:<60} {paginas:>3} página(s)")
            elif paginas == 0:
                resultados.append((ruta_relativa, 0, "VACÍO"))
                print(f"   ⚠ {ruta_relativa:<60} {'0':>3} página(s) [VACÍO]")
            else:
                resultados.append((ruta_relativa, 0, "ERROR"))
                print(f"   ✗ {ruta_relativa:<60} {'ERR':>3} [ERROR LECTURA]")
    
    return resultados


def obtener_rutas_blogs(blogs_seleccionados: List[str] = None) -> Dict[str, str]:
    """
    Obtiene las rutas completas de los blogs a procesar.
    
    Args:
        blogs_seleccionados: Lista de nombres de blogs a procesar (None = todos)
        
    Returns:
        Diccionario {nombre_blog: ruta_completa}
    """
    rutas = {}
    base_path = Path(RUTA_BASE_PUBLICACIONES)
    
    # Procesar blogs estándar
    blogs_a_procesar = blogs_seleccionados if blogs_seleccionados else BLOGS_ESTANDAR
    
    for blog in blogs_a_procesar:
        if blog in BLOGS_ESTANDAR:
            ruta = base_path / blog / "_site"
            if ruta.exists():
                rutas[blog] = str(ruta)
    
    # Si no se especificaron blogs o se incluyeron los especiales
    if not blogs_seleccionados or any(b in ["blog", "teching", "website-achalma"] for b in blogs_seleccionados):
        for nombre, ruta_relativa in BLOGS_WEBSITE_ACHALMA.items():
            if not blogs_seleccionados or nombre in blogs_seleccionados or "website-achalma" in blogs_seleccionados:
                ruta = base_path / ruta_relativa
                if ruta.exists():
                    rutas[f"website-achalma/{nombre}"] = str(ruta)
    
    return rutas


def crear_excel(resultados_por_blog: Dict[str, List[Tuple[str, int, str]]], 
                archivo_salida: str, 
                solo_index: bool):
    """
    Crea un archivo Excel con los resultados.
    
    Args:
        resultados_por_blog: Diccionario {nombre_blog: lista de tuplas}
        archivo_salida: Nombre del archivo Excel de salida
        solo_index: Si se buscaron solo archivos index.pdf
    """
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo de Páginas"
    
    # Estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    blog_font = Font(bold=True, size=11, color="FFFFFF")
    blog_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados principales
    ws['A1'] = "Blog"
    ws['B1'] = "Ruta del Archivo"
    ws['C1'] = "Número de Páginas"
    ws['D1'] = "Estado"
    
    # Aplicar estilos a encabezados
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = header_alignment
        ws[cell].border = thin_border
    
    # Agregar datos
    fila_actual = 2
    total_paginas_global = 0
    total_archivos = 0
    
    for blog, resultados in sorted(resultados_por_blog.items()):
        if not resultados:
            continue
        
        # Fila de encabezado del blog
        ws[f'A{fila_actual}'] = blog.upper()
        ws[f'A{fila_actual}'].font = blog_font
        ws[f'A{fila_actual}'].fill = blog_fill
        ws[f'A{fila_actual}'].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f'A{fila_actual}:D{fila_actual}')
        fila_actual += 1
        
        # Datos del blog
        total_paginas_blog = 0
        for ruta, paginas, estado in resultados:
            ws[f'A{fila_actual}'] = ""  # Blog name (empty for data rows)
            ws[f'B{fila_actual}'] = ruta
            ws[f'C{fila_actual}'] = paginas if estado == "OK" else 0
            ws[f'C{fila_actual}'].alignment = Alignment(horizontal="center")
            ws[f'D{fila_actual}'] = estado
            ws[f'D{fila_actual}'].alignment = Alignment(horizontal="center")
            
            if estado == "OK":
                total_paginas_blog += paginas
            
            fila_actual += 1
            total_archivos += 1
        
        # Subtotal del blog
        ws[f'A{fila_actual}'] = ""
        ws[f'B{fila_actual}'] = f"SUBTOTAL {blog}"
        ws[f'B{fila_actual}'].font = Font(bold=True, italic=True)
        ws[f'C{fila_actual}'] = total_paginas_blog
        ws[f'C{fila_actual}'].font = Font(bold=True, italic=True)
        ws[f'C{fila_actual}'].alignment = Alignment(horizontal="center")
        ws[f'C{fila_actual}'].fill = total_fill
        ws[f'D{fila_actual}'] = f"{len(resultados)} archivos"
        ws[f'D{fila_actual}'].font = Font(italic=True)
        ws[f'D{fila_actual}'].alignment = Alignment(horizontal="center")
        
        total_paginas_global += total_paginas_blog
        fila_actual += 2  # Espacio entre blogs
    
    # Total general
    fila_total = fila_actual
    ws[f'A{fila_total}'] = ""
    ws[f'B{fila_total}'] = "TOTAL GENERAL"
    ws[f'B{fila_total}'].font = total_font
    ws[f'C{fila_total}'] = total_paginas_global
    ws[f'C{fila_total}'].font = total_font
    ws[f'C{fila_total}'].alignment = Alignment(horizontal="center")
    ws[f'C{fila_total}'].fill = total_fill
    ws[f'D{fila_total}'] = f"{total_archivos} archivos"
    ws[f'D{fila_total}'].font = total_font
    ws[f'D{fila_total}'].alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    
    # Agregar metadatos
    ws = wb.create_sheet("Información")
    ws['A1'] = "Información del Reporte"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = "Fecha de generación:"
    ws['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws['A4'] = "Tipo de búsqueda:"
    ws['B4'] = "Solo index.pdf" if solo_index else "Todos los PDFs"
    ws['A5'] = "Total de blogs procesados:"
    ws['B5'] = len(resultados_por_blog)
    ws['A6'] = "Total de archivos:"
    ws['B6'] = total_archivos
    ws['A7'] = "Total de páginas:"
    ws['B7'] = total_paginas_global
    ws['A8'] = "Generado por:"
    ws['B8'] = "Edison Achalma - PDF Page Counter"
    
    # Guardar archivo
    wb.save(archivo_salida)


def listar_blogs_disponibles():
    """Lista todos los blogs disponibles"""
    imprimir_seccion("BLOGS DISPONIBLES")
    
    print("\n📚 Blogs estándar:")
    for i, blog in enumerate(BLOGS_ESTANDAR, 1):
        ruta = Path(RUTA_BASE_PUBLICACIONES) / blog / "_site"
        estado = "✓" if ruta.exists() else "✗"
        print(f"   {i:2d}. {estado} {blog}")
    
    print("\n🌐 Blogs en website-achalma:")
    for nombre, ruta_rel in BLOGS_WEBSITE_ACHALMA.items():
        ruta = Path(RUTA_BASE_PUBLICACIONES) / ruta_rel
        estado = "✓" if ruta.exists() else "✗"
        print(f"       {estado} {nombre}")
    
    print("\n✓ = Disponible | ✗ = No encontrado\n")


def main():
    """Función principal"""
    parser = argparse.ArgumentParser(
        description='Contador de páginas PDF para blogs',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  # Contar todos los blogs (solo index.pdf)
  %(prog)s
  
  # Contar blogs específicos
  %(prog)s -b actus-mercator aequilibria
  
  # Contar todos los PDFs (no solo index.pdf)
  %(prog)s --todos
  
  # Listar blogs disponibles
  %(prog)s --listar
  
  # Archivo de salida personalizado
  %(prog)s -o mi_reporte.xlsx
        """
    )
    
    parser.add_argument(
        '-b', '--blogs',
        nargs='+',
        help='Blogs específicos a procesar (nombres separados por espacios)'
    )
    
    parser.add_argument(
        '-t', '--todos',
        action='store_true',
        help='Buscar todos los archivos PDF (no solo index.pdf)'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Nombre del archivo Excel de salida (se guardará en excel_databases/)'
    )
    
    parser.add_argument(
        '-l', '--listar',
        action='store_true',
        help='Listar todos los blogs disponibles y salir'
    )
    
    args = parser.parse_args()
    
    # Mostrar encabezado
    imprimir_encabezado()
    
    # Si solo quiere listar blogs
    if args.listar:
        listar_blogs_disponibles()
        return
    
    # Verificar que la ruta base existe
    if not Path(RUTA_BASE_PUBLICACIONES).exists():
        print(f"❌ Error: La ruta base no existe: {RUTA_BASE_PUBLICACIONES}")
        print("   Por favor, actualiza RUTA_BASE_PUBLICACIONES en el script.")
        return
    
    # Crear directorio para Excel
    excel_dir = crear_directorio_excel()
    
    # Determinar nombre del archivo de salida
    if args.output:
        archivo_salida = excel_dir / args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tipo = "todos" if args.todos else "index"
        archivo_salida = excel_dir / f"conteo_paginas_{tipo}_{timestamp}.xlsx"
    
    # Obtener rutas de blogs
    rutas_blogs = obtener_rutas_blogs(args.blogs)
    
    if not rutas_blogs:
        print("❌ No se encontraron blogs para procesar.")
        print("   Usa --listar para ver los blogs disponibles.")
        return
    
    # Mostrar configuración
    imprimir_seccion("CONFIGURACIÓN")
    print(f"📁 Ruta base: {RUTA_BASE_PUBLICACIONES}")
    print(f"📊 Modo: {'Todos los PDFs' if args.todos else 'Solo index.pdf'}")
    print(f"📝 Blogs a procesar: {len(rutas_blogs)}")
    print(f"💾 Archivo de salida: {archivo_salida.name}")
    
    # Procesar cada blog
    imprimir_seccion("PROCESANDO BLOGS")
    
    resultados_por_blog = {}
    total_archivos = 0
    total_paginas = 0
    total_errores = 0
    
    for i, (nombre_blog, ruta) in enumerate(rutas_blogs.items(), 1):
        print(f"\n📖 [{i}/{len(rutas_blogs)}] Procesando: {nombre_blog}")
        resultados = buscar_pdfs_en_directorio(ruta, solo_index=not args.todos)
        
        if resultados:
            resultados_por_blog[nombre_blog] = resultados
            archivos_ok = sum(1 for _, _, estado in resultados if estado == "OK")
            paginas_blog = sum(p for _, p, estado in resultados if estado == "OK")
            errores_blog = sum(1 for _, _, estado in resultados if estado == "ERROR")
            
            total_archivos += len(resultados)
            total_paginas += paginas_blog
            total_errores += errores_blog
            
            print(f"   📊 Resumen: {archivos_ok} OK | {errores_blog} errores | {paginas_blog} páginas")
    
    if not resultados_por_blog:
        print("\n⚠️  No se encontraron archivos PDF en ningún blog.")
        return
    
    # Crear reporte Excel
    imprimir_seccion("GENERANDO REPORTE")
    print(f"📝 Creando archivo Excel...")
    crear_excel(resultados_por_blog, str(archivo_salida), solo_index=not args.todos)
    print(f"✅ Archivo creado: {archivo_salida}")
    
    # Mostrar resumen final
    imprimir_resumen(total_archivos, total_paginas, total_errores)
    
    print(f"💡 Tip: El archivo se guardó en: {excel_dir}/")
    print(f"💡 Tip: Usa --listar para ver todos los blogs disponibles\n")


if __name__ == "__main__":
    main()